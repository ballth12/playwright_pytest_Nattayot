import pytest
from playwright.sync_api import Page, expect
import openpyxl
import os
from openpyxl.drawing.image import Image

# กำหนดโฟลเดอร์สำหรับบันทึกภาพหน้าจอในกรณีทดสอบ logout
screenshot_folder = "screenshots/logout01"
os.makedirs(screenshot_folder, exist_ok=True)  # สร้างโฟลเดอร์หากยังไม่มี

def update_excel(test_name, status, screenshot_path=None, copied_text=None):
    """
    ฟังก์ชันสำหรับอัปเดตผลการทดสอบลงในไฟล์ Excel
    - test_name: ชื่อของการทดสอบ
    - status: สถานะการทดสอบ ("passed" หรือ "failed")
    - screenshot_path: ที่อยู่ของไฟล์ภาพหน้าจอ (ถ้ามี)
    - copied_text: ข้อความที่ต้องคัดลอกจากคอลัมน์ที่กำหนด (ใช้เมื่อสถานะเป็น passed)
    """
    try:
        # โหลดไฟล์ Excel
        workbook = openpyxl.load_workbook("testsheet_lab_starter_test.xlsx")
        sheet = workbook["logout01"]  # เลือกชีตที่ใช้บันทึกผลการทดสอบ
        
        # ค้นหาแถวที่ตรงกับชื่อการทดสอบ
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                if cell.value == test_name.split("_")[2]:  # ใช้ตัวเลขรหัสของ test case เพื่อค้นหา
                    row_num = cell.row  # ดึงหมายเลขแถวที่ตรงกัน
                    sheet.cell(row=row_num, column=8, value=status)  # อัปเดตสถานะที่คอลัมน์ H (8)

                    # เพิ่มภาพหน้าจอหากมี
                    if screenshot_path:
                        img = Image(screenshot_path)
                        img.width, img.height = 300, 200  # กำหนดขนาดรูปภาพให้เหมาะสม
                        sheet.add_image(img, f"I{row_num}")  # ใส่รูปในคอลัมน์ I (9)

                    # คัดลอกข้อมูลจากคอลัมน์ F (6) ไปยังคอลัมน์ G (7) ถ้าสถานะเป็น "passed"
                    if status == "passed":
                        copied_from_f = sheet.cell(row=row_num, column=6).value
                        if copied_from_f:
                            sheet.cell(row=row_num, column=7, value=copied_from_f)

                    # บันทึกและปิดไฟล์ Excel
                    workbook.save("testsheet_lab_starter_test.xlsx")
                    workbook.close()
                    return
    except Exception as e:
        print(f"Error updating Excel: {e}")

# ฟังก์ชันทดสอบกระบวนการออกจากระบบ (Logout Process)
def test_logout01_pos01(page: Page):
    try:
        # ไปที่หน้าล็อกอินของระบบ
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
        
        # คลิกปุ่ม "ถอยกลับ" เพื่อกลับไปหน้าหลักของระบบ
        page.click("text=ถอยกลับ", timeout=0)
        
        # คลิกปุ่ม "ออกจากระบบ"
        page.click("text=ออกจากระบบ")
        
        # ตรวจสอบว่ามีปุ่ม "เข้าสู่ระบบ" แสดงอยู่ ซึ่งหมายความว่าการออกจากระบบสำเร็จ
        login_button = page.get_by_role("link", name="เข้าสู่ระบบ", exact=True)
        expect(login_button).to_be_visible()
        
        # รอให้ปุ่ม "เข้าสู่ระบบ" ปรากฏขึ้นก่อนทำการบันทึกภาพหน้าจอ
        screenshot_path = os.path.join(screenshot_folder, "logout01_pos01.png")
        page.screenshot(path=screenshot_path)
        
        # อัปเดตผลการทดสอบลงในไฟล์ Excel
        update_excel("test_logout01_pos01", "passed", screenshot_path)
    except Exception as e:
        print(f"Test failed: {e}")
        
        # หากเกิดข้อผิดพลาด ให้บันทึกภาพหน้าจอเพื่อใช้ในการตรวจสอบ
        screenshot_path = os.path.join(screenshot_folder, "logout01_pos01_failed.png")
        page.screenshot(path=screenshot_path)
        
        # อัปเดตผลลัพธ์ว่า "failed" ลงในไฟล์ Excel
        update_excel("test_logout01_pos01", "failed", screenshot_path)
