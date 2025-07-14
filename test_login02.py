import pytest
from playwright.sync_api import Page, expect, sync_playwright
import openpyxl
import os
from openpyxl.drawing.image import Image


@pytest.fixture(scope="function")
def page():
    with sync_playwright() as p:
        browser = p.firefox.launch(headless=False)  # เปิด browser (headless=False เพื่อให้เห็น UI)
        page = browser.new_page()
        yield page  # ส่ง page object ให้ฟังก์ชันทดสอบใช้งาน
        browser.close()  # ปิด browser หลังจากจบการทดสอบ

# สร้างโฟลเดอร์สำหรับบันทึกภาพหากยังไม่มี
screenshot_folder = "screenshots/login02"
os.makedirs(screenshot_folder, exist_ok=True)

def update_excel(test_name, status, screenshot_path=None, copied_text=None):
    """อัปเดตข้อมูลในไฟล์ Excel และแนบรูปภาพ"""
    try:
        workbook = openpyxl.load_workbook("testsheet_lab_starter_test.xlsx")
        sheet = workbook["login02"]
        
        # ค้นหาแถวที่ตรงกับ test_name
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                if cell.value == test_name.split("_")[2]:  # ตรวจสอบชื่อ testcase
                    row_num = cell.row
                    sheet.cell(row=row_num, column=8, value=status)  # อัปเดตสถานะการทดสอบ

                    # แนบรูปภาพผลการทดสอบ
                    if screenshot_path:
                        img = Image(screenshot_path)
                        img.width, img.height = 300, 200  # กำหนดขนาดรูปภาพ
                        sheet.add_image(img, f"I{row_num}")  # ใส่รูปในคอลัมน์ I (9)

                    # คัดลอกข้อมูลจากคอลัมน์ F ไปยังคอลัมน์ G หากผ่านการทดสอบ
                    if status == "passed":
                        copied_from_f = sheet.cell(row=row_num, column=6).value
                        if copied_from_f:
                            sheet.cell(row=row_num, column=7, value=copied_from_f)

                    workbook.save("testsheet_lab_starter_test.xlsx")
                    workbook.close()  # ปิดไฟล์หลังจากบันทึก
                    return
    except Exception as e:
        print(f"Error updating Excel: {e}")

# ทดสอบการเข้าสู่ระบบด้วยข้อมูลที่ถูกต้อง (กรอกข้อมูลทั้งหมด)
def test_login02_pos01(page: Page):
    try:
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
        page.locator('input[name="f_uid"]').fill("65502100037-8")
        page.locator('input[name="f_pwd"]').fill("0957102316z")
        page.locator('input[name="f_idcard"]').fill("1100401216869")
        page.locator('input[type="submit"][value=" เข้าสู่ระบบ "]').click()
        page.wait_for_selector("text=ยืนยันรหัสผ่านใหม่", timeout=0)
        
        # บันทึกภาพหน้าจอเมื่อผ่านการทดสอบ
        screenshot_path = os.path.join(screenshot_folder, "test_login02_pos01.png")
        page.screenshot(path=screenshot_path)
        copied_text = page.locator("text=ยืนยันรหัสผ่านใหม่").text_content()
        update_excel("test_login02_pos01", "passed", screenshot_path, copied_text)
    except Exception as e:
        print(f"Test failed: {e}")
        screenshot_path = os.path.join(screenshot_folder, "test_login02_pos01_failed.png")
        page.screenshot(path=screenshot_path)
        update_excel("test_login02_pos01", "failed", screenshot_path)

# ทดสอบการเข้าสู่ระบบด้วยรหัสผ่านและหมายเลขบัตรประชาชนที่ไม่ถูกต้อง
def test_login02_neg01(page: Page):
    try:
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
        page.locator('input[name="f_uid"]').fill("65502100037-8")
        page.locator('input[name="f_pwd"]').fill("999999")
        page.locator('input[name="f_idcard"]').fill("99999999999")
        page.locator('input[type="submit"][value=" เข้าสู่ระบบ "]').click()
        
        # ตรวจสอบข้อความ
        error_message_locator = page.locator("text=ก รุ ณ า ป้ อ น ร หั ส ป ร ะ จ ำ ตั ว น ศ .ร หั ส บั ต ร ป ร ะ ช า ช น แ ล ะ ร หั ส ผ่ า น ใ ห้ ถู ก ต้ อ ง")
        expect(error_message_locator).to_be_visible()
        
        # ตรวจสอบสีของข้อความ
        error_message_element = error_message_locator.element_handle()
        color = page.evaluate('(element) => getComputedStyle(element).color', error_message_element)
        assert color == 'rgb(0, 0, 255)'
        
        # บันทึกภาพหน้าจอเมื่อผ่านการทดสอบ
        screenshot_path = os.path.join(screenshot_folder, "test_login02_neg01.png")
        page.screenshot(path=screenshot_path)
        update_excel("test_login02_neg01", "passed", screenshot_path)
    except Exception as e:
        print(f"Test failed: {e}")
        screenshot_path = os.path.join(screenshot_folder, "test_login02_neg01_failed.png")
        page.screenshot(path=screenshot_path)
        update_excel("test_login02_neg01", "failed", screenshot_path)

# ทดสอบการเข้าสู่ระบบด้วยหมายเลขบัตรประชาชนที่ไม่ถูกต้อง
def test_login02_neg02(page: Page):
    try:
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
        page.locator('input[name="f_uid"]').fill("65502100037-8")
        page.locator('input[name="f_pwd"]').fill("0957102316z")
        page.locator('input[name="f_idcard"]').fill("99999999999")
        page.locator('input[type="submit"][value=" เข้าสู่ระบบ "]').click()

        # ตรวจสอบข้อความ
        error_message_locator = page.locator("text=ก รุ ณ า ป้ อ น ร หั ส ป ร ะ จ ำ ตั ว น ศ .ร หั ส บั ต ร ป ร ะ ช า ช น แ ล ะ ร หั ส ผ่ า น ใ ห้ ถู ก ต้ อ ง")
        expect(error_message_locator).to_be_visible()

        # ตรวจสอบสีของข้อความ
        error_message_element = error_message_locator.element_handle()
        color = page.evaluate('(element) => getComputedStyle(element).color', error_message_element)
        assert color == 'rgb(0, 0, 255)'

        # บันทึกภาพหน้าจอเมื่อผ่านการทดสอบ
        screenshot_path = os.path.join(screenshot_folder, "test_login02_neg02.png")
        page.screenshot(path=screenshot_path)
        update_excel("test_login02_neg02", "passed", screenshot_path)
    except Exception as e:
        print(f"Test failed: {e}")
        screenshot_path = os.path.join(screenshot_folder, "test_login02_neg02_failed.png")
        page.screenshot(path=screenshot_path)
        update_excel("test_login02_neg02", "failed", screenshot_path)

# ทดสอบการเข้าสู่ระบบด้วย รหัสประจำตัว รหัสผ่าน และหมายเลขบัตรประชาชนที่ไม่ถูกต้อง
def test_login02_neg03(page: Page):
    try:
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
        page.locator('input[name="f_uid"]').fill("55555555")
        page.locator('input[name="f_pwd"]').fill("555555")
        page.locator('input[name="f_idcard"]').fill("99999999999")
        page.locator('input[type="submit"][value=" เข้าสู่ระบบ "]').click()

        # ตรวจสอบข้อความ
        error_message_locator = page.locator("text=กรุณาป้อนรหัสประจำตัวและรหัสผ่านให้ถูกต้อง")
        expect(error_message_locator).to_be_visible()

        # ตรวจสอบสีของข้อความ
        error_message_element = error_message_locator.element_handle()
        color = page.evaluate('(element) => getComputedStyle(element).color', error_message_element)
        assert color == 'rgb(0, 0, 255)'

        # บันทึกภาพหน้าจอเมื่อผ่านการทดสอบ
        screenshot_path = os.path.join(screenshot_folder, "test_login02_neg03.png")
        page.screenshot(path=screenshot_path)
        update_excel("test_login02_neg03", "passed", screenshot_path)
    except Exception as e:
        print(f"Test failed: {e}")
        screenshot_path = os.path.join(screenshot_folder, "test_login02_neg03_failed.png")
        page.screenshot(path=screenshot_path)
        update_excel("test_login02_neg03", "failed", screenshot_path)